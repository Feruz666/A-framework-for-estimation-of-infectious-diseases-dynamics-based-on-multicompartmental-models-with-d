from cProfile import label
import random
import scipy.integrate
import numpy as np
import matplotlib.pyplot as plt
from scipy import interpolate
import xlsxwriter
import pandas as pd
import openpyxl as xl
import statistics
from scipy.interpolate import interp1d, make_interp_spline
np.set_printoptions(suppress=True)

def averageOfList(num):
    sumOfNumbers = 0
    for t in num:
        sumOfNumbers = sumOfNumbers + t

    avg = sumOfNumbers / len(num)
    return avg

def odes(x , t):
    #assign each ODE to a vector element
    S   = x[0]
    Ea  = x[1]
    E   = x[2]
    Ia  = x[3]
    I   = x[4]
    R   = x[5]

    #define each ODE
    dsdt = (
        (-beta_transmission_rate[0]*I*S)/n
    ) - (
        ( rIa_infect_of_asympt[0]* beta_transmission_rate[0]*Ia*S)/n
    ) - (
        ( rE_infect_of_pre_sympt[0]* beta_transmission_rate[0]*E*S)/n
    )
    
    deadt = (
        ( beta_transmission_rate[0]*I*S)/n
    ) - (
        (1/ lmda_a_exposed_period[0])*Ea
    )

    dedt = (
        ( rE_infect_of_pre_sympt[0]* beta_transmission_rate[0]*Ea*S)/n
    ) + (
        (1/ lmda_a_exposed_period[0])*(1- p_a_probab_of_asymp_infect[0])*Ea
    ) - (
        (1/ lmda_pre_sympt_period[0])*E
    )

    diadt = (
        ( rIa_infect_of_asympt[0]* beta_transmission_rate[0]*Ia*S)/n
    ) + (
        (1/ lmda_a_exposed_period[0])* p_a_probab_of_asymp_infect[0]*Ea
    ) - (
        (1/ gamma_infect_period[0]) * Ia
    )

    didt = (
        (1/ gamma_infect_period[0]) * Ia
        ) + (
            (1/ lmda_pre_sympt_period[0]) * E
        ) - (
            (1/ gamma_infect_period[0]) * I
        )
    
    drdt = (1/ gamma_infect_period[0])*(I + Ia)

    return [dsdt, deadt, dedt, diadt, didt, drdt]


wb = xl.load_workbook('./writer.xlsx')
ws = wb.active
ws.delete_rows(1, ws.max_row+1) # for entire sheet

workbook = xlsxwriter.Workbook('writer.xlsx')
worksheet = workbook.add_worksheet()



#initial conditions
x0 = [99990, 0, 0, 0, 10, 0]
i = 0
final_arr = []
days = 200
points = 250
t = np.linspace(0, days, points)

while i < 200:
    beta_transmission_rate = np.random.normal(0.225, 0.04, 1)
    rIa_infect_of_asympt  = np.random.normal(0.68, 0.06, 1)
    rE_infect_of_pre_sympt  = np.random.normal(0.675, 0.05, 1)
    lmda_a_exposed_period  = np.random.normal(4, 0.666, 1)
    lmda_pre_sympt_period  = np.random.normal(3.5, 0.5, 1)
    p_a_probab_of_asymp_infect  = np.random.normal(0.475, 0.05, 1)
    gamma_infect_period  = np.random.normal(8, 1.33, 1)
    n = 100000
    x = scipy.integrate.odeint(odes, x0, t)
    final_arr.append(x[:,4])
    i+=1

for row_num, row_data in enumerate(final_arr):
    for col_num, col_data in enumerate(row_data):
        worksheet.write(row_num, col_num, col_data)
        
workbook.close()

df2 = pd.read_excel ('writer.xlsx')
df_list = df2.to_numpy()
tr_dflist = df_list.transpose()
tr_dflist = tr_dflist.tolist()
m_list = []
start_50_l = []
end_50_l = []
start_75_l = []
end_75_l = []
start_95_l = []
end_95_l = []
i=1
for i in range(100):
# for i in range(70,len(final_arr), 1):
    
    m = averageOfList(tr_dflist[i])
    start_50, end_50 = np.percentile(tr_dflist[i], 25), np.percentile(tr_dflist[i], 75)
    start_75, end_75 = np.percentile(tr_dflist[i], 12.5), np.percentile(tr_dflist[i], 87.5)
    start_95, end_95 = np.percentile(tr_dflist[i], 2.5), np.percentile(tr_dflist[i], 97.5)

    if m <= 0.:
        m = 0
    if start_50 <= 0.:
        start_50 = 0
    if end_50 <= 0.:
        end_50 = 0
    if start_75 <= 0.:
        start_75 = 0
    if end_75 <= 0.:
        end_75 = 0
    if start_95 <= 0.:
        start_95 = 0
    if  end_95 <= 0.:
        end_95 = 0

    
    m_list.append(m)
    start_50_l.append(start_50)
    end_50_l.append(end_50)
    start_75_l.append(start_75)
    end_75_l.append(end_75)
    start_95_l.append(start_95)
    end_95_l.append(end_95)


#Mean
x = np.arange(len(m_list))
M_BSpline = interpolate.interp1d(x, m_list, kind="cubic")
xm_new = np.arange(x[0], x[-1], 0.1)
ym_new = M_BSpline(xm_new)

for i in range(len(ym_new)):
    if ym_new[i] <= 0:
        ym_new[i] = 0

#Start_50
S50_BSpline = interpolate.interp1d(x, start_50_l, kind="cubic")
xs50_new = np.arange(x[0], x[-1], 0.1)
ys50_new = S50_BSpline(xs50_new)
for i in range(len(ys50_new)):
    if ys50_new[i] <= 0:
        ys50_new[i] = 0

#End_50
E50BSpline = interpolate.interp1d(x, end_50_l, kind="cubic")
xe50_new = np.arange(x[0], x[-1], 0.1)
ye50_new = E50BSpline(xe50_new)
for i in range(len(ye50_new)):
    if ye50_new[i] <= 0:
        ye50_new[i] = 0

#Start_75
S75BSpline = interpolate.interp1d(x, start_75_l, kind="cubic")
xs75_new = np.arange(x[0], x[-1], 0.1)
ys75_new = S75BSpline(xs75_new)
for i in range(len(ys75_new)):
    if ys75_new[i] <= 0:
        ys75_new[i] = 0

#End_75
E75BSpline = interpolate.interp1d(x, end_75_l, kind="cubic")
xe75_new = np.arange(x[0], x[-1], 0.1)
ye75_new = E75BSpline(xe75_new)
for i in range(len(ye75_new)):
    if ye75_new[i] <= 0:
        ye75_new[i] = 0

#Start_95
S95BSpline = interpolate.interp1d(x, start_95_l, kind="cubic")
xs95_new = np.arange(x[0], x[-1], 0.1)
ys95_new = S95BSpline(xs95_new)
for i in range(len(ys95_new)):
    if ys95_new[i] <= 0:
        ys95_new[i] = 0

#End_95
E95BSpline = interpolate.interp1d(x, end_95_l, kind="cubic")
xe95_new = np.arange(x[0], x[-1], 0.1)
ye95_new = E95BSpline(xe95_new)
for i in range(len(ye95_new)):
    if ye95_new[i] <= 0:
        ye95_new[i] = 0




dataset = pd.read_csv("new_cases.csv")

col_list = ["date", "Russia"]
date = dataset['date'].tolist()
russia = dataset['Russia'].tolist()



date_data = []
russia_data = []
for i in range(0, 210):
    date_data.append(date[i])
    russia_data.append(russia[i])


n = 10
date_data = date_data[n:]
russia_data = russia_data[n:]

x = np.arange(len(russia_data))
M_BSpline = make_interp_spline(x, russia_data)
rf_xm_new = np.arange(x[0], x[-1], 0.1)
rf_ym_new = M_BSpline(rf_xm_new)



fig = plt.figure()
ax = fig.add_subplot(1,1,1)

x_ticks = np.arange(0, 1000, 10)

ax.set_xticks(x_ticks)



ax.grid()
ax.plot(xm_new, ym_new, 'r-', label = "Mean")
ax.plot(xs50_new,ys50_new, 'y-', label = "50% R", markersize=4)
ax.plot(xe50_new,ye50_new, 'y-', label = "50% L", markersize=4)
ax.plot(xs75_new,ys75_new, 'g-', label = "75% R", markersize=4)
ax.plot(xe75_new,ye75_new, 'g-', label = "75% L", markersize=4)
ax.plot(xs95_new,ys95_new, 'b-', label = "95% R", markersize=4)
ax.plot(xe95_new,ye95_new, 'b-', label = "95% L", markersize=4)
plt.plot(rf_xm_new, rf_ym_new, 'k-', label = "Real data", markersize=4)
ax.legend()
plt.show()
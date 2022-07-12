import random
import scipy.integrate
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import FormatStrFormatter
import matplotlib.ticker as tc
from scipy import interpolate
import xlsxwriter
import pandas as pd
import openpyxl as xl

np.set_printoptions(suppress=True)


def odes(x , t):
    #assign each ODE to a vector element
    S1   = x[0]
    S2   = x[1]
    Ea1  = x[2]
    E1   = x[3]
    Ea2  = x[4]
    E2   = x[5]
    Ia1  = x[6]
    Ia2  = x[7]
    I1   = x[8]
    I2   = x[9]
    R    = x[10]

    #define each ODE
    ds1dt = (
        (-beta_transmission_rate11*I1*S1)/n
    ) - (
        (rIa_infect_of_asympt*beta_transmission_rate11*Ia1*S1)/n
    ) - (
        (rE_infect_of_pre_sympt*beta_transmission_rate11*E1*S1)/n
    ) - (
        (-beta_transmission_rate12*I2*S2)/n
    ) - (
        (rIa_infect_of_asympt*beta_transmission_rate12*Ia2*S1)/n
    ) - (
        (rE_infect_of_pre_sympt*beta_transmission_rate12*E2*S1)/n
    )

    ds2dt = (
        (-beta_transmission_rate22*I2*S2)/n
    ) - (
        (rIa_infect_of_asympt*beta_transmission_rate22*Ia2*S2)/n
    ) - (
        (rE_infect_of_pre_sympt*beta_transmission_rate22*E2*S2)/n
    ) - (
        (-beta_transmission_rate21*I1*S2)/n
    ) - (
        (rIa_infect_of_asympt*beta_transmission_rate21*Ia1*S2)/n
    ) - (
        (rE_infect_of_pre_sympt*beta_transmission_rate21*E1*S2)/n
    )
    
    dea1dt = (
        (beta_transmission_rate11*I1*S1)/n
    ) + (
        (beta_transmission_rate12*I2*S1)/n
    ) - (
        (1/lmda_a_exposed_period)*Ea1
    )

    de1dt = (
        (rE_infect_of_pre_sympt*beta_transmission_rate11*E1*S1)/n
    ) + (
        (rE_infect_of_pre_sympt*beta_transmission_rate12*E2*S1)/n
    ) + (
        (1/lmda_a_exposed_period)*(1-p_a_probab_of_asymp_infect)*Ea1
    ) - (
        (1/lmda_pre_sympt_period)*E1
    )

    dea2dt = (
        (beta_transmission_rate22*I2*S2)/n
    ) + (
        (beta_transmission_rate21*I1*S2)/n
    ) - (
        (1/lmda_a_exposed_period)*Ea2
    )

    de2dt = (
        (rE_infect_of_pre_sympt*beta_transmission_rate22*E2*S2)/n
    ) + (
        (rE_infect_of_pre_sympt*beta_transmission_rate21*E1*S2)/n
    ) + (
        (1/lmda_a_exposed_period)*(1-p_a_probab_of_asymp_infect)*Ea2
    ) - (
        (1/lmda_pre_sympt_period)*E2
    )

    dia1dt = (
        (rIa_infect_of_asympt*beta_transmission_rate11*Ia1*S1)/n
    ) + (
        (rIa_infect_of_asympt*beta_transmission_rate12*Ia2*S1)/n
    ) + (
        (1/lmda_a_exposed_period)*p_a_probab_of_asymp_infect*Ea1
    ) - (
        (1/gamma_infect_period1) * Ia1
    )

    dia2dt = (
        (rIa_infect_of_asympt*beta_transmission_rate22*Ia2*S2)/n
    ) + (
        (rIa_infect_of_asympt*beta_transmission_rate21*Ia1*S2)/n
    ) + (
        (1/lmda_a_exposed_period)*p_a_probab_of_asymp_infect*Ea2
    ) - (
        (1/gamma_infect_period2) * Ia2
    )

    di1dt = (
        (1/gamma_infect_period1) * Ia1
    ) + (
        (1/lmda_pre_sympt_period) * E1
    ) - (
        (1/gamma_infect_period1) * I1
    )

    di2dt = (
        (1/gamma_infect_period2) * Ia2
    ) + (
        (1/lmda_pre_sympt_period) * E2
    ) - (
        (1/gamma_infect_period2) * I2
    )
    
    drdt = (
        (1/gamma_infect_period1)*(I1 + Ia1)
    ) + (
        (1/gamma_infect_period2)*(I2 + Ia2)
    )

    return [ds1dt, ds2dt, dea1dt, de1dt, dea2dt, de2dt, dia1dt, dia2dt, di1dt, di2dt, drdt]


wb = xl.load_workbook('./writer.xlsx')
ws = wb.active
ws.delete_rows(1, ws.max_row+1) # for entire sheet

workbook = xlsxwriter.Workbook('writer.xlsx')
worksheet = workbook.add_worksheet()



#initial conditions
x0 = [14999, 14999, 0, 0, 0, 0, 0, 0, 1, 1, 0]
t = np.linspace(0, 300, 400)


i = 0

final_arr = []

beta_transmission_rate11_arange = np.arange(0.08, 0.3, 0.01, float)
beta_transmission_rate_arange = np.arange(0.08, 0.3, 0.01, float)

beta_transmission_rate12_arange = np.arange(0.1, 0.46, 0.01, float)
beta_transmission_rate21_arange = np.arange(0.1, 0.46, 0.01, float)
beta_transmission_rate22_arange = np.arange(0.1, 0.46, 0.01, float)

rIa_infect_of_asympt_arange = np.arange(0.5, 0.9, 0.1, float)   
rE_infect_of_pre_sympt_arange = np.arange(0.5, 0.85, 0.01, float)      
lmda_a_exposed_period_arange = np.arange(2, 6, 1, int)
lmda_pre_sympt_period_arange = np.arange(2, 6, 1, int)
p_a_probab_of_asymp_infect_arange = np.arange(0.3, 0.65, 0.01, float)

gamma_infect_period1_arange = np.arange(0.07, 0.19, 0.01, float)
gamma_infect_period2_arange = np.arange(0.11, 0.28, 0.01, float)



while i < 200:
    beta_transmission_rate11    = random.choice([0.08, 0.3])
    beta_transmission_rate      = random.choice([0.08, 0.3])

    beta_transmission_rate12 = random.choice([0.1, 0.46])
    beta_transmission_rate21 = random.choice([0.1, 0.46])
    beta_transmission_rate22 = random.choice([0.1, 0.46])

    rIa_infect_of_asympt        = random.choice([0.5, 0.9])
    rE_infect_of_pre_sympt      = random.choice([0.5, 0.85])
    lmda_a_exposed_period       = random.choice([2,6])
    lmda_pre_sympt_period       = random.choice([2,6])
    p_a_probab_of_asymp_infect  = random.choice([0.3, 0.65])

    gamma_infect_period1         = random.choice([0.07, 0.19])
    gamma_infect_period2         = random.choice([0.11, 0.28])

    n = 30000

    x = scipy.integrate.odeint(odes, x0, t)
    final_arr.append(x[:,9])
    i+=1

for row_num, row_data in enumerate(final_arr):
    for col_num, col_data in enumerate(row_data):
        worksheet.write(row_num, col_num, col_data)

workbook.close()



df2 = pd.read_excel ('writer.xlsx')
df_list = df2.to_numpy()
tr_dflist = df_list.transpose()
tr_dflist = tr_dflist.tolist()
m_list = []
start_50_l = []
end_50_l = []
start_75_l = []
end_75_l = []
start_95_l = []
end_95_l = []
i=1
for i in range(400):
# for i in range(70,len(final_arr), 1):
    
    m = np.mean(tr_dflist[i])
    start_50, end_50 = np.percentile(tr_dflist[i], 25), np.percentile(tr_dflist[i], 75)
    start_75, end_75 = np.percentile(tr_dflist[i], 12.5), np.percentile(tr_dflist[i], 87.5)
    start_95, end_95 = np.percentile(tr_dflist[i], 2.5), np.percentile(tr_dflist[i], 97.5)

    if m <= 0.:
        m = 0
    if start_50 <= 0.:
        start_50 = 0
    if end_50 <= 0.:
        end_50 = 0
    if start_75 <= 0.:
        start_75 = 0
    if end_75 <= 0.:
        end_75 = 0
    if start_95 <= 0.:
        start_95 = 0
    if  end_95 <= 0.:
        end_95 = 0

    
    m_list.append(m)
    start_50_l.append(start_50)
    end_50_l.append(end_50)
    start_75_l.append(start_75)
    end_75_l.append(end_75)
    start_95_l.append(start_95)
    end_95_l.append(end_95)


#Mean
x = np.arange(len(m_list))
M_BSpline = interpolate.interp1d(x, m_list, kind="cubic")
xm_new = np.arange(x[0], x[-1], 0.1)
ym_new = M_BSpline(xm_new)

for i in range(len(ym_new)):
    if ym_new[i] <= 0:
        ym_new[i] = 0

#Start_50
S50_BSpline = interpolate.interp1d(x, start_50_l, kind="cubic")
xs50_new = np.arange(x[0], x[-1], 0.1)
ys50_new = S50_BSpline(xs50_new)
for i in range(len(ys50_new)):
    if ys50_new[i] <= 0:
        ys50_new[i] = 0

#End_50
E50BSpline = interpolate.interp1d(x, end_50_l, kind="cubic")
xe50_new = np.arange(x[0], x[-1], 0.1)
ye50_new = E50BSpline(xe50_new)
for i in range(len(ye50_new)):
    if ye50_new[i] <= 0:
        ye50_new[i] = 0

#Start_75
S75BSpline = interpolate.interp1d(x, start_75_l, kind="cubic")
xs75_new = np.arange(x[0], x[-1], 0.1)
ys75_new = S75BSpline(xs75_new)
for i in range(len(ys75_new)):
    if ys75_new[i] <= 0:
        ys75_new[i] = 0

#End_75
E75BSpline = interpolate.interp1d(x, end_75_l, kind="cubic")
xe75_new = np.arange(x[0], x[-1], 0.1)
ye75_new = E75BSpline(xe75_new)
for i in range(len(ye75_new)):
    if ye75_new[i] <= 0:
        ye75_new[i] = 0

#Start_95
S95BSpline = interpolate.interp1d(x, start_95_l, kind="cubic")
xs95_new = np.arange(x[0], x[-1], 0.1)
ys95_new = S95BSpline(xs95_new)
for i in range(len(ys95_new)):
    if ys95_new[i] <= 0:
        ys95_new[i] = 0

#End_95
E95BSpline = interpolate.interp1d(x, end_95_l, kind="cubic")
xe95_new = np.arange(x[0], x[-1], 0.1)
ye95_new = E95BSpline(xe95_new)
for i in range(len(ye95_new)):
    if ye95_new[i] <= 0:
        ye95_new[i] = 0


fig = plt.figure()
ax = fig.add_subplot(1,1,1)

x_ticks = np.arange(0, 1000, 10)

ax.set_xticks(x_ticks)



ax.grid()
ax.plot(xm_new, ym_new, 'r-', label = "Mean")
ax.plot(xs50_new,ys50_new, 'y-', label = "50% R", markersize=4)
ax.plot(xe50_new,ye50_new, 'y-', label = "50% L", markersize=4)
ax.plot(xs75_new,ys75_new, 'g-', label = "75% R", markersize=4)
ax.plot(xe75_new,ye75_new, 'g-', label = "75% L", markersize=4)
ax.plot(xs95_new,ys95_new, 'b-', label = "95% R", markersize=4)
ax.plot(xe95_new,ye95_new, 'b-', label = "95% L", markersize=4)
ax.legend()
plt.show()